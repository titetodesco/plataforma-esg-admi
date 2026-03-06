import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def df_from_query(conn, sql: str, params=()):
    cur = conn.execute(sql, params)
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description] if cur.description else []
    return pd.DataFrame(rows, columns=cols)

def autosize(ws, max_width=70):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 80) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), max_width)

def write_df(ws, df):
    ws.append(list(df.columns))
    for r in df.itertuples(index=False, name=None):
        ws.append(list(r))
    ws.freeze_panes = "A2"
    autosize(ws)


def build_validacap_setup(questionario_id: str, cfg: pd.DataFrame, pt: pd.DataFrame, ptop: pd.DataFrame,
                          ind: pd.DataFrame, fr: pd.DataFrame, rt: pd.DataFrame, re: pd.DataFrame,
                          calc_active: pd.DataFrame) -> pd.DataFrame:
    rows = []

    def add(regra: str, ok: bool, detalhe: str, qtd=None):
        rows.append({
            "regra": regra,
            "status": "OK" if ok else "ERRO",
            "detalhe": detalhe,
            "qtd": "" if qtd is None else qtd,
        })

    add("Questionário informado", bool(str(questionario_id).strip()), f"questionario_id={questionario_id}")

    add(
        "CONFIG_ORGANIZACOES possui 1 registro",
        len(cfg) == 1,
        f"Encontrado(s): {len(cfg)}",
        len(cfg),
    )

    add(
        "PESOS_TEMA não vazio",
        len(pt) > 0,
        f"Encontrado(s): {len(pt)}",
        len(pt),
    )

    add(
        "PESOS_TOPICO não vazio",
        len(ptop) > 0,
        f"Encontrado(s): {len(ptop)}",
        len(ptop),
    )

    add(
        "INDICADORES_SETUP não vazio",
        len(ind) > 0,
        f"Encontrado(s): {len(ind)}",
        len(ind),
    )

    if not ind.empty and "indicador_id" in ind.columns:
        dup_ind = int(ind.duplicated(subset=["indicador_id"], keep=False).sum())
        add(
            "INDICADORES_SETUP sem indicador_id duplicado",
            dup_ind == 0,
            "Sem duplicados" if dup_ind == 0 else "Há indicador_id duplicado",
            dup_ind,
        )
    else:
        add("INDICADORES_SETUP contém coluna indicador_id", False, "Coluna ausente ou tabela vazia")

    if not fr.empty and {"indicador_id", "nivel"}.issubset(set(fr.columns)):
        dup_fr = int(fr.duplicated(subset=["indicador_id", "nivel"], keep=False).sum())
        add(
            "FAIXA_REFERENCIA sem duplicidade (indicador_id,nivel)",
            dup_fr == 0,
            "Sem duplicados" if dup_fr == 0 else "Há duplicidade por indicador/nível",
            dup_fr,
        )
    else:
        add("FAIXA_REFERENCIA contém colunas mínimas", False, "Necessário: indicador_id e nivel")

    calc_ids = []
    if not calc_active.empty and "indicador_id" in calc_active.columns:
        calc_ids = sorted(calc_active["indicador_id"].astype(str).tolist())

    if calc_ids and not fr.empty and {"indicador_id", "nivel"}.issubset(set(fr.columns)):
        incompletos = []
        fr2 = fr.copy()
        fr2["nivel"] = pd.to_numeric(fr2["nivel"], errors="coerce")
        for ind_id in calc_ids:
            niveis = set(fr2.loc[fr2["indicador_id"].astype(str) == str(ind_id), "nivel"].dropna().astype(int).tolist())
            if niveis != {1, 2, 3, 4, 5}:
                incompletos.append(ind_id)
        add(
            "Indicadores calculados ativos com 5 faixas",
            len(incompletos) == 0,
            "Todos completos" if len(incompletos) == 0 else f"Incompletos: {', '.join(incompletos[:10])}",
            len(incompletos),
        )
    else:
        add(
            "Indicadores calculados ativos com 5 faixas",
            len(calc_ids) == 0,
            "Sem calculados ativos" if len(calc_ids) == 0 else "Não foi possível validar faixas",
            len(calc_ids),
        )

    add(
        "RECOM_TEMA não vazio",
        len(rt) > 0,
        f"Encontrado(s): {len(rt)}",
        len(rt),
    )

    add(
        "RECOM_EIXO não vazio",
        len(re) > 0,
        f"Encontrado(s): {len(re)}",
        len(re),
    )

    return pd.DataFrame(rows, columns=["regra", "status", "detalhe", "qtd"])

def export_setup_xlsx(conn, questionario_id: str) -> bytes:
    conn.sync()  # 1 sync apenas aqui (você já aplicou isso ✅)

    wb = Workbook()
    wb.remove(wb.active)

    # Puxa o questionário (para montar CONFIG_ORGANIZACOES)
    q = df_from_query(conn, """
        SELECT setor, porte, regiao, status, observacao
        FROM questionario
        WHERE questionario_id=?
    """, (questionario_id,))
    if q.empty:
        raise ValueError(f"Questionario_id '{questionario_id}' não encontrado.")

    # CONFIG_ORGANIZACOES
    cfg = q.rename(columns={"setor":"SETOR","porte":"PORTE","regiao":"REGIAO","observacao":"OBSERVACAO"}).copy()
    cfg["ATIVO"] = cfg["status"].apply(lambda s: 0 if str(s).upper() == "ARCHIVED" else 1)
    cfg = cfg[["SETOR","PORTE","REGIAO","ATIVO","OBSERVACAO"]]
    write_df(wb.create_sheet("CONFIG_ORGANIZACOES"), cfg)

    # PESOS_TEMA
    pt = df_from_query(conn, """
        SELECT questionario_id, tema_id, peso_tema
        FROM peso_tema
        WHERE questionario_id=?
        ORDER BY tema_id
    """, (questionario_id,))
    if pt.empty:
        pt = pd.DataFrame(columns=["questionario_id","tema_id","peso_tema"])
    write_df(wb.create_sheet("PESOS_TEMA"), pt)

    # PESOS_TOPICO
    ptop = df_from_query(conn, """
        SELECT questionario_id, topico_id, peso_topico
        FROM peso_topico
        WHERE questionario_id=?
        ORDER BY topico_id
    """, (questionario_id,))
    if ptop.empty:
        ptop = pd.DataFrame(columns=["questionario_id","topico_id","peso_topico"])
    write_df(wb.create_sheet("PESOS_TOPICO"), ptop)

    # INDICADORES_SETUP
    ind = df_from_query(conn, """
        SELECT questionario_id, indicador_id, ativo, peso_indicador
        FROM indicador_config
        WHERE questionario_id=?
        ORDER BY indicador_id
    """, (questionario_id,))
    if ind.empty:
        ind = pd.DataFrame(columns=["questionario_id","indicador_id","ativo","peso_indicador"])
    write_df(wb.create_sheet("INDICADORES_SETUP"), ind)

    # FAIXA_REFERENCIA
    fr_raw = df_from_query(conn, """
        SELECT questionario_id, indicador_id, nivel, tipo_regra, valor_min, valor_max, valor_exato, rotulo
        FROM faixa_referencia
        WHERE questionario_id=?
        ORDER BY indicador_id, nivel
    """, (questionario_id,))

    calc_active = df_from_query(conn, """
        SELECT ic.indicador_id
        FROM indicador_config ic
        JOIN indicador i ON i.indicador_id = ic.indicador_id
        WHERE ic.questionario_id=?
          AND ic.ativo=1
          AND i.tipo_indicador='CALCULADO'
        ORDER BY ic.indicador_id
    """, (questionario_id,))

    if calc_active.empty:
        fr = fr_raw if not fr_raw.empty else pd.DataFrame(columns=["questionario_id","indicador_id","nivel","tipo_regra","valor_min","valor_max","valor_exato","rotulo"])
    else:
        indicadores_calc = calc_active["indicador_id"].tolist()
        expected_idx = pd.MultiIndex.from_product(
            [indicadores_calc, [1, 2, 3, 4, 5]],
            names=["indicador_id", "nivel"],
        )

        if fr_raw.empty:
            fr = expected_idx.to_frame(index=False)
            fr["questionario_id"] = questionario_id
            fr["tipo_regra"] = "INTERVALO"
            fr["valor_min"] = None
            fr["valor_max"] = None
            fr["valor_exato"] = None
            fr["rotulo"] = ""
            fr = fr[["questionario_id","indicador_id","nivel","tipo_regra","valor_min","valor_max","valor_exato","rotulo"]]
        else:
            fr_calc = fr_raw[fr_raw["indicador_id"].isin(indicadores_calc)].copy()
            fr_calc = fr_calc.set_index(["indicador_id", "nivel"]).reindex(expected_idx).reset_index()
            fr_calc["questionario_id"] = fr_calc["questionario_id"].fillna(questionario_id)
            fr_calc["tipo_regra"] = fr_calc["tipo_regra"].fillna("INTERVALO")
            fr_calc["rotulo"] = fr_calc["rotulo"].fillna("")
            fr = fr_calc[["questionario_id","indicador_id","nivel","tipo_regra","valor_min","valor_max","valor_exato","rotulo"]]

    write_df(wb.create_sheet("FAIXA_REFERENCIA"), fr)

    # RECOM_TEMA (somente tabela editável)
    rt = df_from_query(conn, """
        SELECT questionario_id, tema_id, nivel, recomendacao
        FROM recomendacao_tema
        WHERE questionario_id=?
        ORDER BY tema_id, nivel
    """, (questionario_id,))
    if rt.empty:
        rt = pd.DataFrame(columns=["questionario_id","tema_id","nivel","recomendacao"])

    write_df(wb.create_sheet("RECOM_TEMA"), rt)

    # RECOM_EIXO (somente tabela editável)
    re = df_from_query(conn, """
        SELECT questionario_id, eixo_id, nivel, recomendacao
        FROM recomendacao_eixo
        WHERE questionario_id=?
        ORDER BY eixo_id, nivel
    """, (questionario_id,))
    if re.empty:
        re = pd.DataFrame(columns=["questionario_id","eixo_id","nivel","recomendacao"])

    write_df(wb.create_sheet("RECOM_EIXO"), re)

    # VALIDACAP_SETUP (regras de consistência do setup exportado)
    valid = build_validacap_setup(questionario_id, cfg, pt, ptop, ind, fr, rt, re, calc_active)
    write_df(wb.create_sheet("VALIDACAP_SETUP"), valid)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
