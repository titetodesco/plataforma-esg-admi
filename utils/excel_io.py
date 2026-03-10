import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _df_from_query(conn, sql: str, params=()):
    cur = conn.execute(sql, params)
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description] if cur.description else []
    return pd.DataFrame(rows, columns=cols)


def _autosize(ws, max_width=70):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 120) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), max_width)


def _write_df(ws, df):
    ws.append(list(df.columns))
    for r in df.itertuples(index=False, name=None):
        ws.append(list(r))
    ws.freeze_panes = "A2"
    _autosize(ws)


def export_macrobase_xlsx(conn) -> bytes:
    conn.sync()

    wb = Workbook()
    wb.remove(wb.active)

    eixos = _df_from_query(
        conn,
        """
        SELECT eixo_id AS EIXO_ID, codigo AS CODIGO, nome AS NOME, descricao AS DESCRICAO, peso_default AS PESO_DEFAULT
        FROM eixo
        ORDER BY eixo_id
        """,
    )
    _write_df(wb.create_sheet("EIXOS"), eixos)

    temas = _df_from_query(
        conn,
        """
        SELECT tema_id AS TEMA_ID, eixo_id AS EIXO_ID, codigo AS CODIGO, nome AS NOME, descricao AS DESCRICAO, peso_default AS PESO_DEFAULT
        FROM tema
        ORDER BY tema_id
        """,
    )
    _write_df(wb.create_sheet("TEMAS"), temas)

    topicos = _df_from_query(
        conn,
        """
        SELECT topico_id AS TOPICO_ID, tema_id AS TEMA_ID, codigo AS CODIGO, nome AS NOME, descricao AS DESCRICAO, peso_default AS PESO_DEFAULT
        FROM topico
        ORDER BY topico_id
        """,
    )
    _write_df(wb.create_sheet("TOPICOS"), topicos)

    indicadores = _df_from_query(
        conn,
        """
        SELECT
            indicador_id AS INDICADOR_ID,
            topico_id AS TOPICO_ID,
            codigo AS CODIGO,
            nome AS NOME,
            descricao AS DESCRICAO,
            tipo_indicador AS TIPO_INDICADOR,
            psr_tipo AS PSR_TIPO,
            formula AS FORMULA,
            unidade_resultado AS UNIDADE_RESULTADO,
            peso_default AS PESO_DEFAULT
        FROM indicador
        ORDER BY indicador_id
        """,
    )
    _write_df(wb.create_sheet("INDICADORES"), indicadores)

    variaveis = _df_from_query(
        conn,
        """
        SELECT
            variavel_id AS VARIAVEL_ID,
            codigo AS CODIGO,
            pergunta_texto AS PERGUNTA_TEXTO,
            descricao AS DESCRICAO,
            tipo_resposta AS TIPO_RESPOSTA,
            unidade_entrada AS UNIDADE_ENTRADA,
            observacoes AS OBSERVACOES
        FROM variavel
        ORDER BY variavel_id
        """,
    )
    _write_df(wb.create_sheet("VARIAVEIS"), variaveis)

    variavel_opcoes = _df_from_query(
        conn,
        """
        SELECT variavel_id AS VARIAVEL_ID, ordem AS ORDEM, texto_opcao AS TEXTO_OPCAO, score_1a5 AS SCORE_1A5
        FROM variavel_opcao
        ORDER BY variavel_id, ordem
        """,
    )
    _write_df(wb.create_sheet("VARIAVEL_OPCOES"), variavel_opcoes)

    indicador_variavel = _df_from_query(
        conn,
        """
        SELECT indicador_id AS INDICADOR_ID, variavel_id AS VARIAVEL_ID, papel AS PAPEL, obrigatoria AS OBRIGATORIA, peso AS PESO
        FROM indicador_variavel
        ORDER BY indicador_id, variavel_id
        """,
    )
    _write_df(wb.create_sheet("INDICADOR_VARIAVEL"), indicador_variavel)

    recom_tema_default = _df_from_query(
        conn,
        """
        SELECT tema_id AS TEMA_ID, nivel AS NIVEL, recomendacao AS RECOMENDACAO
        FROM recomendacao_tema_default
        ORDER BY tema_id, nivel
        """,
    )
    _write_df(wb.create_sheet("RECOMENDACAO_TEMA_DEFAULT"), recom_tema_default)

    recom_eixo_default = _df_from_query(
        conn,
        """
        SELECT eixo_id AS EIXO_ID, nivel AS NIVEL, recomendacao AS RECOMENDACAO
        FROM recomendacao_eixo_default
        ORDER BY eixo_id, nivel
        """,
    )
    _write_df(wb.create_sheet("RECOMENDACAO_EIXO_DEFAULT"), recom_eixo_default)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def load_macrobase(file):
    xls = pd.ExcelFile(file)

    data = {
        "EIXOS": pd.read_excel(file, sheet_name="EIXOS"),
        "TEMAS": pd.read_excel(file, sheet_name="TEMAS"),
        "TOPICOS": pd.read_excel(file, sheet_name="TOPICOS"),
        "INDICADORES": pd.read_excel(file, sheet_name="INDICADORES"),
        "VARIAVEIS": pd.read_excel(file, sheet_name="VARIAVEIS"),
        "VARIAVEL_OPCOES": pd.read_excel(file, sheet_name="VARIAVEL_OPCOES"),
        "INDICADOR_VARIAVEL": pd.read_excel(file, sheet_name="INDICADOR_VARIAVEL"),
    }

    optional_aliases = {
        "RECOMENDACAO_TEMA_DEFAULT": ["RECOMENDACAO_TEMA_DEFAULT", "RECOM_TEMA_DEFAULT"],
        "RECOMENDACAO_EIXO_DEFAULT": ["RECOMENDACAO_EIXO_DEFAULT", "RECOM_EIXO_DEFAULT"],
    }

    available = set(xls.sheet_names)
    for key, aliases in optional_aliases.items():
        chosen = next((name for name in aliases if name in available), None)
        if chosen:
            data[key] = pd.read_excel(file, sheet_name=chosen)

    return data
